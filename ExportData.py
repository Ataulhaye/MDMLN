import csv
from datetime import datetime
from string import ascii_uppercase

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

from ExelSettings import ExelSettings
from ExportEntity import ExportEntity


class ExportData:
    def create_and_write_csv(self, data, sheet_name, title="results", delimiter=","):
        csv_data = self.prepare_data_matrix(data)
        file_name = self.get_file_name(extension=".csv", sheet_name=sheet_name)
        with open(file_name, "w", newline="") as file:
            for ed in csv_data:
                wr = csv.writer(file, delimiter=delimiter, quoting=csv.QUOTE_ALL)
                wr.writerow(ed)

    def create_and_write_datasheet(
        self,
        data,
        sheet_name,
        notes,
        title="results",
        transpose=False,
        single_label=False,
    ):
        # matrix = data
        matrix = self.prepare_data_matrix(data)

        if transpose is True:
            matrix = np.transpose(matrix)

        # to create a new blank Workbook object
        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = title

        sett = ExelSettings()

        col_widths = self.calculate_column_width(matrix)

        cell_positions = self.calculate_cell_positions(matrix)

        for i, row in enumerate(matrix[0]):
            ws.column_dimensions[cell_positions[0][i][0]].width = col_widths[i]

        for row in matrix:
            if type(row) != list:
                row = list(row)
            ws.append(row)

        self.set_font_significant_result(cell_positions, matrix, ws, transpose)

        self.set_header_font(cell_positions, matrix, ws, col_widths, sett)

        self.set_first_column_font(cell_positions, matrix, ws, col_widths, sett)

        if not single_label:
            self.merge_strategy_cells(transpose, matrix, ws)
        increment = 2
        if single_label:
            increment = 1
        self.fill_color(cell_positions, matrix, ws, sett, increment)

        # step incrment x to 2 to ..
        # ws.merge_cells(start_row=x, start_column=1, end_row=x, end_column=1)

        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        # for cell in rows:
        # cell.fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")

        if transpose:
            self.set_2nd_column_font(
                cell_positions, matrix, ws, col_widths, sett, transpose
            )
        else:
            self.set_2nd_row_font(
                cell_positions, matrix, ws, col_widths, sett, transpose
            )

        sheet_name = self.get_file_name(extension=".xlsx", sheet_name=sheet_name)

        for note in notes:
            ws.append(note)

        self.set_note_font(notes, matrix, ws, sett)

        # Save File
        wb.save(sheet_name)

    def set_note_font(self, notes, matrix, ws, sett):
        note_cell_coord = None
        for col in ws.iter_cols(min_row=1, max_col=1, max_row=(matrix.shape[0] + 6)):
            for cell in col:
                if cell.internal_value == notes[-1][0]:
                    note_cell_coord = cell.coordinate
                    break

        if note_cell_coord is not None:
            note_cell = ws[note_cell_coord]
            note_cell.font = Font(name=sett.header_family, sz=sett.header_font)

    def create_note(self, t_config):
        notes = []
        for i in range(5):
            notes.append([])
        # notes.append([])
        # notes.append([])
        # notes.append([])
        # notes.append([])
        notes.append([str(t_config)])
        return notes

    def fill_color(self, cell_positions, matrix, ws, setting: ExelSettings, increment):
        i = 0
        for row in matrix:
            j = 0
            for cont in row:
                cell = ws[cell_positions[i][j]]
                if i == 0:
                    cell.fill = PatternFill(
                        start_color=setting.header_cell_fill_color,
                        end_color=setting.header_cell_fill_color,
                        fill_type=setting.cell_fill_type,
                    )
                else:
                    cell.fill = PatternFill(
                        start_color=setting.cell_fill_color,
                        end_color=setting.cell_fill_color,
                        fill_type=setting.cell_fill_type,
                    )
                j = j + 1
            if i > 0 and i % increment == 0:
                i = i + increment
            i = i + 1
            if len(matrix) <= i:
                break

    @staticmethod
    def merge_strategy_cells(transpose, matrix, ws):
        x = 2
        for row in matrix:
            if transpose:
                sc = 1
                ec = 1
                rs = x
                re = x + 1
            else:
                sc = x
                ec = x + 1
                rs = 1
                re = 1
            ws.merge_cells(start_row=rs, start_column=sc, end_row=re, end_column=ec)
            if x % 2 == 0:
                x = x + 2
            if re >= len(matrix):
                break

    def set_first_column_font(
        self, cell_position, matrix, ws, col_widths, setting: ExelSettings
    ):
        for j, row in enumerate(matrix):
            cell = ws[cell_position[j][0]]
            cell.font = Font(name=setting.header_family, sz=setting.header_font)

        self.repair_first_column_width(cell_position, ws, col_widths, setting)

    def set_2nd_column_font(
        self, cell_position, matrix, ws, col_widths, setting: ExelSettings, transpose
    ):
        for j, row in enumerate(matrix):
            cell = ws[cell_position[j][1]]
            cell.font = Font(name=setting.header_family, sz=setting.header_font)

        self.repair_column_width(cell_position, ws, col_widths, setting, 1)

    def set_2nd_row_font(
        self, cell_position, matrix, ws, col_widths, setting: ExelSettings, transpose
    ):
        for j, row in enumerate(matrix[1]):
            cell = ws[cell_position[1][j]]
            cell.font = Font(name=setting.header_family, sz=setting.header_font)

        self.repair_row_width(cell_position, ws, col_widths, matrix)

    @staticmethod
    def repair_first_column_width(cell_position, ws, col_widths, setting: ExelSettings):
        ws.column_dimensions[cell_position[0][0][0]].width = col_widths[0] + (
            (setting.header_font - setting.default_font) * 2
        )

    @staticmethod
    def repair_column_width(
        cell_position, ws, col_widths, setting: ExelSettings, column: int
    ):
        ws.column_dimensions[cell_position[0][column][0]].width = col_widths[column] + (
            (setting.header_font - setting.default_font) * 2
        )

    @staticmethod
    def repair_row_width(cell_position, ws, col_widths, matrix):
        for i, row in enumerate(matrix[0]):
            if i > 0:
                ws.column_dimensions[cell_position[0][i][0]].width = col_widths[i]

    def set_header_font(
        self, cell_position, matrix, ws, col_widths, setting: ExelSettings
    ):
        for i, cell in enumerate(matrix[0]):
            cell = ws[cell_position[0][i]]
            cell.font = Font(name=setting.header_family, sz=setting.header_font)
            self.repair_header_cell_width(cell_position, ws, col_widths, i, setting)

    @staticmethod
    def repair_header_cell_width(
        cell_position, ws, col_widths, i, setting: ExelSettings
    ):
        ws.column_dimensions[cell_position[0][i][0]].width = col_widths[i] + (
            (setting.header_font - setting.default_font)
        )

    @staticmethod
    def set_font_significant_result(cell_position, matrix, ws, transpose):
        k = 0
        if transpose:
            k = 1
        for i, row in enumerate(matrix):
            for j, val in enumerate(row):
                if "Not" not in val and j > k and i > 0:
                    cell = ws[cell_position[i][j]]
                    cell.font = Font(bold=True)

    @staticmethod
    def calculate_cell_positions(matrix):
        cell_names = []
        for i, row in enumerate(matrix):
            row_names = []
            for j, val in enumerate(row):
                row_names.append(f"{ascii_uppercase[j]}{i+1}")
            cell_names.append(row_names)
        return cell_names

    @staticmethod
    def calculate_column_width(matrix):
        max_column_widths = [0] * len(matrix[0])
        for row in matrix:
            for i, cell in enumerate(row):
                cell_length = len(cell)
                if max_column_widths[i] < cell_length:
                    max_column_widths[i] = cell_length
        return max_column_widths

    @staticmethod
    def get_file_name(extension, sheet_name):
        dt = datetime.now().strftime("%d-%m-%Y_%H-%M-%S_%f")
        sheet_name = f"{sheet_name}_{dt}{extension}"
        return sheet_name

    def prepare_data_matrix(self, data: list[ExportEntity]):
        for scn in data:
            if scn.sub_column_name is None:
                scn.sub_column_name = "None"

        data.sort(key=lambda x: x.sub_column_name)
        matrix = [["Strategy"], ["Classifier"]]

        for i, export_entity in enumerate(data):
            sub_col = f"{export_entity.sub_column_name}#{export_entity.column_name}"
            col = f"{export_entity.column_name}#{export_entity.sub_column_name}"
            if col not in matrix[1]:
                matrix[1].append(col)
            if sub_col not in matrix[0]:
                matrix[0].append(sub_col)
            row_col_index = self.get_index(matrix, export_entity.row_name)

            col_index = matrix[1].index(col)

            if row_col_index is not None:
                row_index = row_col_index[0]
                row = matrix[row_index]
                size = len(row) - 1
                if size < col_index:
                    for i in range(col_index - size):
                        row.append(None)

                row[col_index] = f"{export_entity.result[0]} {export_entity.result[1]}"

                matrix[row_index] = row

            else:
                insert_row = [None] * (col_index + 1)
                insert_row[0] = export_entity.row_name
                insert_row[col_index] = (
                    f"{export_entity.result[0]} {export_entity.result[1]}"
                )

                matrix.append(insert_row)

        for i in range(2):
            for j, cell in enumerate(matrix[i]):
                matrix[i][j] = cell.split("#")[0]

        return matrix

    @staticmethod
    def init_matrix(rows, columns):
        matrix = []
        for r in range(rows):
            matrix.append([None] * columns)
        return matrix

    @staticmethod
    def matrix_dimensions(data: list[ExportEntity]):
        row_data, col_data = [], []
        for export_entity in data:
            col_data.append(
                f"{export_entity.sub_column_name}-{export_entity.column_name}"
            )
            row_data.append(export_entity.row_name)
        rows = len(set(row_data)) + 2
        columns = len(set(col_data)) + 1
        return rows, columns

    # returns row, col index of a given value
    @staticmethod
    def get_index(matrix, value):
        for i, row in enumerate(matrix):
            if value in row:
                return i, row.index(value)
